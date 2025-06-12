import os
from multiprocessing import Pool, cpu_count, freeze_support
from tqdm import tqdm
import openpyxl

def convert_sheet_to_csv(sheet_info):
    file_path, sheet_name = sheet_info
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    sheet = wb[sheet_name]

    csv_file = f"{os.path.splitext(file_path)[0]}_{sheet_name}.csv"
    with open(csv_file, "w", encoding="utf-8") as f:
        for row in sheet.iter_rows(values_only=True):
            f.write(",".join(['' if v is None else str(v) for v in row]) + "\n")
    return sheet_name

def convert_all_sheets_parallel(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet_names = wb.sheetnames
    tasks = [(file_path, sheet) for sheet in sheet_names]

    print(f"🔧 Converting {len(sheet_names)} sheets using {min(cpu_count(), len(sheet_names))} workers...")

    with Pool(processes=min(cpu_count(), len(sheet_names))) as pool:
        for _ in tqdm(pool.imap_unordered(convert_sheet_to_csv, tasks), total=len(tasks)):
            pass

if __name__ == "__main__":
    freeze_support()
    file_path = r"C:\Users\harsh\Downloads\Crime_Data_from_2020_to_Present-Test.xlsx"
    convert_all_sheets_parallel(file_path)


