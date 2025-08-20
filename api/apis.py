from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from api.models import Projects, EDAPlot, EDAFormat
from django.conf import settings
from django.shortcuts import redirect
from django.http import HttpResponse
import os
import json
from google_auth_oauthlib.flow import Flow
from django.utils import timezone
from .models import APILog
from rest_framework import serializers
from django.core.paginator import Paginator, EmptyPage
from django.db.models import Q
import pandas as pd
from .serializers import EDAPlotSerializer
from django.shortcuts import get_object_or_404
from datetime import datetime
from rest_framework.parsers import JSONParser
# Excel and chart imports
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.chart.reference import Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from io import BytesIO

# Write new API views here

class ExampleNewAPI(APIView):
    def get(self, request):
        return Response({'message': 'This is a new API!'}, status=status.HTTP_200_OK)

class APILogSerializer(serializers.ModelSerializer):
    class Meta:
        model = APILog
        fields = '__all__'

class LoggingAPI(APIView):
    def get(self, request):
        level = request.query_params.get('level')
        user_id = request.query_params.get('user_id')
        project_id = request.query_params.get('project_id')
        file_name = request.query_params.get('file_name')
        sheet_name = request.query_params.get('sheet_name')
        page = int(request.query_params.get('page', 1))
        page_size = 15

        logs = APILog.objects.all()

        # Filter out specific endpoints
        excluded_endpoints = [
            'getCommits', 'getCommitsArray', 'auth', 'projectss', 
            'get-shared-projects', 'save-pivot', 'save-plots', 'fetch-pivot-plots', 'update-sheet-data', 'upload', 'fetch_scripts', 'project-details', 'fetch-report-pivot'
        ]
        
        # Create Q objects to exclude these endpoints
        exclude_conditions = Q()
        for endpoint in excluded_endpoints:
            exclude_conditions |= Q(endpoint__icontains=endpoint)
        
        logs = logs.exclude(exclude_conditions)

        if level == 'user' and user_id:
            logs = logs.filter(user_id=user_id)
        elif level == 'project' and project_id:
            logs = logs.filter(project_id=project_id)
        elif level == 'sheet' and project_id and file_name and sheet_name:
            logs = logs.filter(project_id=project_id, file_name=file_name, sheet_name=sheet_name)
        else:
            return Response({'error': 'Invalid or missing parameters.'}, status=400)

        paginator = Paginator(logs, page_size)
        try:
            page_obj = paginator.page(page)
        except EmptyPage:
            return Response({'logs': [], 'page': page, 'total_pages': paginator.num_pages, 'total_logs': paginator.count}, status=200)

        serializer = APILogSerializer(page_obj.object_list, many=True)
        return Response({
            'logs': serializer.data,
            'page': page,
            'total_pages': paginator.num_pages,
            'total_logs': paginator.count
        }, status=200) 

class SaveEDAPlot(APIView):
    """
    API to save EDA plots with comprehensive validation and error handling
    """
    parser_classes = [JSONParser]

    def post(self, request):
        try:
            # Extract data from request
            user_id = request.data.get('user_id')
            project_id = request.data.get('project_id')
            plot_name = request.data.get('plot_name', f'EDA Plot {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
            plot_type = request.data.get('plot_type', 'custom')
            
            # Data source information
            file_type = request.data.get('file_type')
            file_name = request.data.get('file_name')
            sheet_name = request.data.get('sheet_name')
            
            # Plot configuration and data
            plot_config = request.data.get('plot_config', {})
            chart_data = request.data.get('chart_data', {})
            chart_options = request.data.get('chart_options', {})
            
            # Handle both old and new Y-axis formats
            # Check if plot_config contains yAxes (new format) or yAxis (old format)
            if 'yAxes' in plot_config and plot_config['yAxes']:
                # New format with multiple Y-axes
                y_axes = plot_config['yAxes']
                y_axis = y_axes[0] if y_axes else None  # Use first Y-axis for backward compatibility
            elif 'yAxis' in plot_config and plot_config['yAxis']:
                # Old format with single Y-axis
                y_axis = plot_config['yAxis']
                y_axes = [y_axis]  # Convert to array for new format
            else:
                # No Y-axis specified
                y_axis = None
                y_axes = []
            
            # Update plot_config to ensure both formats are available
            plot_config['yAxis'] = y_axis
            plot_config['yAxes'] = y_axes
            
            # EDA-specific metadata
            eda_analysis_type = request.data.get('eda_analysis_type')
            
            # Handle columns_analyzed to include all X and Y axes
            x_axes = plot_config.get('xAxes', [])
            columns_analyzed = list(set(x_axes + y_axes))  # Remove duplicates
            
            data_summary = request.data.get('data_summary', {})
            # Update data_summary to include both formats
            data_summary.update({
                'x_axes': x_axes,
                'y_axis': y_axis,
                'y_axes': y_axes,
                'chart_type': plot_config.get('chartType', 'bar'),
                'aggregation_method': plot_config.get('aggregationMethod', 'sum'),
                'date_grouping': plot_config.get('dateGrouping', 'raw')
            })
            
            # Additional metadata
            description = request.data.get('description', '')
            tags = request.data.get('tags', [])

            # Validate required fields
            if not all([user_id, project_id]):
                missing_fields = []
                if not user_id: missing_fields.append('user_id')
                if not project_id: missing_fields.append('project_id')
                return Response({
                    'error': f'Missing required fields: {", ".join(missing_fields)}'
                }, status=400)

            # Validate plot_type
            valid_plot_types = [choice[0] for choice in EDAPlot.PLOT_TYPES]
            if plot_type not in valid_plot_types:
                return Response({
                    'error': f'Invalid plot_type. Must be one of: {", ".join(valid_plot_types)}'
                }, status=400)

            # Get user and project
            try:
                from .models import User
                user = User.objects.get(id=user_id)
                project = Projects.objects.get(id=project_id)
            except User.DoesNotExist:
                return Response({'error': 'User not found'}, status=404)
            except Projects.DoesNotExist:
                return Response({'error': 'Project not found'}, status=404)

            # Check if user has access to this project
            if project.user != user:
                return Response({
                    'error': 'Access denied. You don\'t have permission to save plots for this project.'
                }, status=403)

            # Create or update EDA plot
            eda_plot, created = EDAPlot.objects.update_or_create(
                user=user,
                project=project,
                plot_name=plot_name,
                defaults={
                    'plot_type': plot_type,
                    'file_type': file_type,
                    'file_name': file_name,
                    'sheet_name': sheet_name,
                    'plot_config': plot_config,
                    'chart_data': chart_data,
                    'chart_options': chart_options,
                    'eda_analysis_type': eda_analysis_type,
                    'columns_analyzed': columns_analyzed,
                    'data_summary': data_summary,
                    'description': description,
                    'tags': tags
                }
            )

            # Log the action
            ip = request.META.get('REMOTE_ADDR')
            from .log_utils import log_user_action
            log_user_action(user, "save_eda_plot", details=f"EDA plot '{plot_name}' saved successfully", ip_address=ip)

            return Response({
                'message': 'EDA plot saved successfully',
                'plot_id': eda_plot.id,
                'plot_name': eda_plot.plot_name,
                'created': created,
                'updated_at': eda_plot.updated_at
            }, status=201 if created else 200)

        except Exception as e:
            return Response({'error': str(e)}, status=500)

class FetchEDAPlots(APIView):
    """
    API to fetch EDA plots with filtering and pagination
    """
    parser_classes = [JSONParser]

    def post(self, request):
        try:
            # Extract data from request
            user_id = request.data.get('user_id')
            project_id = request.data.get('project_id')
            plot_type = request.data.get('plot_type')
            eda_analysis_type = request.data.get('eda_analysis_type')
            file_type = request.data.get('file_type')
            file_name = request.data.get('file_name')
            sheet_name = request.data.get('sheet_name')
            tags = request.data.get('tags', [])
            
            # Pagination
            page = int(request.data.get('page', 1))
            page_size = int(request.data.get('page_size', 15))

            # Validate required fields
            if not all([user_id, project_id]):
                missing_fields = []
                if not user_id: missing_fields.append('user_id')
                if not project_id: missing_fields.append('project_id')
                return Response({
                    'error': f'Missing required fields: {", ".join(missing_fields)}'
                }, status=400)

            # Get user and project
            try:
                from .models import User
                user = User.objects.get(id=user_id)
                project = Projects.objects.get(id=project_id)
            except User.DoesNotExist:
                return Response({'error': 'User not found'}, status=404)
            except Projects.DoesNotExist:
                return Response({'error': 'Project not found'}, status=404)

            # Check if user has access to this project
            if project.user != user:
                return Response({
                    'error': 'Access denied. You don\'t have permission to view plots for this project.'
                }, status=403)

            # Build filter for plots
            plot_filter = {
                'project': project
            }
            
            # Apply optional filters
            if plot_type:
                plot_filter['plot_type'] = plot_type
            if eda_analysis_type:
                plot_filter['eda_analysis_type'] = eda_analysis_type
            if file_type:
                plot_filter['file_type'] = file_type
            if file_name:
                plot_filter['file_name'] = file_name
            if sheet_name:
                plot_filter['sheet_name'] = sheet_name

            # Fetch EDA plots with filters
            eda_plots = EDAPlot.objects.filter(**plot_filter)

            # Filter by tags if provided
            if tags:
                for tag in tags:
                    eda_plots = eda_plots.filter(tags__contains=[tag])

            # Apply pagination
            paginator = Paginator(eda_plots, page_size)
            try:
                page_obj = paginator.page(page)
            except EmptyPage:
                return Response({
                    'plots': [],
                    'page': page,
                    'total_pages': paginator.num_pages,
                    'total_plots': paginator.count
                }, status=200)

            # Serialize the plots
            serializer = EDAPlotSerializer(page_obj.object_list, many=True)
            
            return Response({
                'plots': serializer.data,
                'page': page,
                'total_pages': paginator.num_pages,
                'total_plots': paginator.count
            }, status=200)

        except Exception as e:
            return Response({'error': str(e)}, status=500)

class UpdateEDAPlot(APIView):
    """
    API to update existing EDA plots
    """
    parser_classes = [JSONParser]

    def put(self, request, plot_id):
        try:
            # Extract data from request
            user_id = request.data.get('user_id')
            
            # Get the plot
            try:
                eda_plot = EDAPlot.objects.get(id=plot_id)
            except EDAPlot.DoesNotExist:
                return Response({'error': 'EDA plot not found'}, status=404)

            # Check if user has permission to update this plot
            # Convert user_id to integer for comparison
            try:
                user_id_int = int(user_id)
            except (ValueError, TypeError):
                return Response({
                    'error': 'Invalid user_id format'
                }, status=400)
            
            if eda_plot.user.id != user_id_int:
                return Response({
                    'error': 'Access denied. You don\'t have permission to update this plot.',
                    'debug_info': {
                        'plot_user_id': eda_plot.user.id,
                        'request_user_id': user_id_int,
                        'plot_id': plot_id
                    }
                }, status=403)

            # Update fields
            update_fields = [
                'plot_name', 'plot_type', 'file_type', 'file_name', 'sheet_name',
                'chart_data', 'chart_options', 'eda_analysis_type',
                'description', 'tags'
            ]
            
            for field in update_fields:
                if field in request.data:
                    setattr(eda_plot, field, request.data[field])
            
            # Handle plot_config specially to ensure Y-axes compatibility
            if 'plot_config' in request.data:
                plot_config = request.data['plot_config']
                
                # Handle both old and new Y-axis formats
                if 'yAxes' in plot_config and plot_config['yAxes']:
                    # New format with multiple Y-axes
                    y_axes = plot_config['yAxes']
                    y_axis = y_axes[0] if y_axes else None
                elif 'yAxis' in plot_config and plot_config['yAxis']:
                    # Old format with single Y-axis
                    y_axis = plot_config['yAxis']
                    y_axes = [y_axis]
                else:
                    # No Y-axis specified
                    y_axis = None
                    y_axes = []
                
                # Update plot_config to ensure both formats are available
                plot_config['yAxis'] = y_axis
                plot_config['yAxes'] = y_axes
                
                # Update columns_analyzed and data_summary
                x_axes = plot_config.get('xAxes', [])
                columns_analyzed = list(set(x_axes + y_axes))
                
                data_summary = eda_plot.data_summary or {}
                data_summary.update({
                    'x_axes': x_axes,
                    'y_axis': y_axis,
                    'y_axes': y_axes,
                    'chart_type': plot_config.get('chartType', 'bar'),
                    'aggregation_method': plot_config.get('aggregationMethod', 'sum'),
                    'date_grouping': plot_config.get('dateGrouping', 'raw')
                })
                
                eda_plot.plot_config = plot_config
                eda_plot.columns_analyzed = columns_analyzed
                eda_plot.data_summary = data_summary

            eda_plot.save()

            # Log the action
            ip = request.META.get('REMOTE_ADDR')
            from .log_utils import log_user_action
            log_user_action(eda_plot.user, "update_eda_plot", details=f"EDA plot '{eda_plot.plot_name}' updated", ip_address=ip)

            serializer = EDAPlotSerializer(eda_plot)
            return Response({
                'message': 'EDA plot updated successfully',
                'plot': serializer.data
            }, status=200)

        except Exception as e:
            return Response({'error': str(e)}, status=500)

class DeleteEDAPlot(APIView):
    """
    API to delete EDA plots
    """
    parser_classes = [JSONParser]

    def delete(self, request, plot_id):
        try:
            # Extract data from query parameters (DELETE requests don't have body)
            user_id = request.query_params.get('user_id')
            
            if not user_id:
                return Response({
                    'error': 'Missing required parameter: user_id'
                }, status=400)
            
            # Get the plot
            try:
                eda_plot = EDAPlot.objects.get(id=plot_id)
            except EDAPlot.DoesNotExist:
                return Response({'error': 'EDA plot not found'}, status=404)

            # Check if user has permission to delete this plot
            if eda_plot.user.id != int(user_id):
                return Response({
                    'error': 'Access denied. You don\'t have permission to delete this plot.'
                }, status=403)

            plot_name = eda_plot.plot_name
            eda_plot.delete()

            # Log the action
            ip = request.META.get('REMOTE_ADDR')
            from .log_utils import log_user_action
            from .models import User
            user = User.objects.get(id=user_id)
            log_user_action(user, "delete_eda_plot", details=f"EDA plot '{plot_name}' deleted", ip_address=ip)

            return Response({
                'message': 'EDA plot deleted successfully'
            }, status=200)

        except Exception as e:
            return Response({'error': str(e)}, status=500)

class GetEDAPlotDetails(APIView):
    """
    API to get detailed information about a specific EDA plot
    """
    parser_classes = [JSONParser]

    def post(self, request):
        try:
            # Extract data from request
            user_id = request.data.get('user_id')
            plot_id = request.data.get('plot_id')
            
            if not all([user_id, plot_id]):
                missing_fields = []
                if not user_id: missing_fields.append('user_id')
                if not plot_id: missing_fields.append('plot_id')
                return Response({
                    'error': f'Missing required fields: {", ".join(missing_fields)}'
                }, status=400)

            # Get the plot
            try:
                eda_plot = EDAPlot.objects.get(id=plot_id)
            except EDAPlot.DoesNotExist:
                return Response({'error': 'EDA plot not found'}, status=404)

            # Check if user has permission to view this plot
            # Convert user_id to integer for comparison
            try:
                user_id_int = int(user_id)
            except (ValueError, TypeError):
                return Response({
                    'error': 'Invalid user_id format'
                }, status=400)
            
            if eda_plot.user.id != user_id_int:
                return Response({
                    'error': 'Access denied. You don\'t have permission to view this plot.'
                }, status=403)

            serializer = EDAPlotSerializer(eda_plot)
            return Response({
                'plot': serializer.data
            }, status=200)

        except Exception as e:
            return Response({'error': str(e)}, status=500) 

class GetEnhancedPlotData(APIView):
    """
    API to get enhanced plot data with support for multiple Y-axes
    """
    parser_classes = [JSONParser]

    def post(self, request):
        try:
            # Extract data from request
            user_id = request.data.get('user_id')
            project_id = request.data.get('project_id')
            plot_id = request.data.get('plot_id')
            
            # Validate required fields
            if not all([user_id, project_id, plot_id]):
                missing_fields = []
                if not user_id: missing_fields.append('user_id')
                if not project_id: missing_fields.append('project_id')
                if not plot_id: missing_fields.append('plot_id')
                return Response({
                    'error': f'Missing required fields: {", ".join(missing_fields)}'
                }, status=400)

            # Get user and project
            try:
                from .models import User
                user = User.objects.get(id=user_id)
                project = Projects.objects.get(id=project_id)
            except User.DoesNotExist:
                return Response({'error': 'User not found'}, status=404)
            except Projects.DoesNotExist:
                return Response({'error': 'Project not found'}, status=404)

            # Check if user has access to this project
            if project.user != user:
                return Response({
                    'error': 'Access denied. You don\'t have permission to view plots for this project.'
                }, status=403)

            # Get the plot
            try:
                eda_plot = EDAPlot.objects.get(id=plot_id, project=project)
            except EDAPlot.DoesNotExist:
                return Response({'error': 'EDA plot not found'}, status=404)

            # Serialize the plot with enhanced data
            serializer = EDAPlotSerializer(eda_plot)
            plot_data = serializer.data
            
            # Add enhanced information
            enhanced_data = {
                'plot': plot_data,
                'x_axes': eda_plot.get_x_axes(),
                'y_axes': eda_plot.get_y_axes(),
                'y_axis': eda_plot.get_y_axes()[0] if eda_plot.get_y_axes() else None,
                'chart_type': eda_plot.get_chart_type(),
                'aggregation_method': eda_plot.get_aggregation_method(),
                'date_grouping': eda_plot.get_date_grouping(),
                'has_multiple_y_axes': len(eda_plot.get_y_axes()) > 1,
                'has_multiple_x_axes': len(eda_plot.get_x_axes()) > 1
            }
            
            return Response(enhanced_data, status=200)

        except Exception as e:
            return Response({'error': str(e)}, status=500)

# EDA Format Management APIs

class SaveEDAFormat(APIView):
    """
    API to save EDA format/flow for reuse
    """
    parser_classes = [JSONParser]

    def post(self, request):
        try:
            # Extract data from request
            user_id = request.data.get('user_id')
            format_name = request.data.get('format_name')
            format_type = request.data.get('format_type', 'single_plot')
            
            # Format configuration
            format_config = request.data.get('format_config', {})
            required_columns = request.data.get('required_columns', [])
            optional_columns = request.data.get('optional_columns', [])
            column_patterns = request.data.get('column_patterns', {})
            
            # Sample data structure
            sample_data_structure = request.data.get('sample_data_structure', {})
            
            # Metadata
            description = request.data.get('description', '')
            tags = request.data.get('tags', [])
            category = request.data.get('category', '')

            # Validate required fields
            if not all([user_id, format_name]):
                missing_fields = []
                if not user_id: missing_fields.append('user_id')
                if not format_name: missing_fields.append('format_name')
                return Response({
                    'error': f'Missing required fields: {", ".join(missing_fields)}'
                }, status=400)

            # Get user
            try:
                from .models import User
                user = User.objects.get(id=user_id)
            except User.DoesNotExist:
                return Response({'error': 'User not found'}, status=404)

            # Create or update EDA format
            eda_format, created = EDAFormat.objects.update_or_create(
                user=user,
                format_name=format_name,
                defaults={
                    'format_type': format_type,
                    'format_config': format_config,
                    'required_columns': required_columns,
                    'optional_columns': optional_columns,
                    'column_patterns': column_patterns,
                    'sample_data_structure': sample_data_structure,
                    'description': description,
                    'tags': tags,
                    'category': category
                }
            )

            # Log the action
            ip = request.META.get('REMOTE_ADDR')
            from .log_utils import log_user_action
            log_user_action(user, "save_eda_format", details=f"EDA format '{format_name}' saved successfully", ip_address=ip)

            return Response({
                'message': 'EDA format saved successfully',
                'format_id': eda_format.id,
                'format_name': eda_format.format_name,
                'created': created,
                'updated_at': eda_format.updated_at
            }, status=201 if created else 200)

        except Exception as e:
            return Response({'error': str(e)}, status=500)

class FetchEDAFormats(APIView):
    """
    API to fetch EDA formats with filtering and pagination
    """
    parser_classes = [JSONParser]

    def post(self, request):
        try:
            # Extract data from request
            user_id = request.data.get('user_id')
            format_type = request.data.get('format_type')
            category = request.data.get('category')
            tags = request.data.get('tags', [])
            
            # Pagination
            page = int(request.data.get('page', 1))
            page_size = int(request.data.get('page_size', 15))

            # Validate required fields
            if not user_id:
                return Response({
                    'error': 'Missing required field: user_id'
                }, status=400)

            # Get user
            try:
                from .models import User
                user = User.objects.get(id=user_id)
            except User.DoesNotExist:
                return Response({'error': 'User not found'}, status=404)

            # Build filter for formats
            format_filter = {
                'user': user,
                'is_active': True
            }
            
            # Apply optional filters
            if format_type:
                format_filter['format_type'] = format_type
            if category:
                format_filter['category'] = category

            # Fetch EDA formats with filters
            eda_formats = EDAFormat.objects.filter(**format_filter)

            # Filter by tags if provided
            if tags:
                for tag in tags:
                    eda_formats = eda_formats.filter(tags__contains=[tag])

            # Apply pagination
            paginator = Paginator(eda_formats, page_size)
            try:
                page_obj = paginator.page(page)
            except EmptyPage:
                return Response({
                    'formats': [],
                    'page': page,
                    'total_pages': paginator.num_pages,
                    'total_formats': paginator.count
                }, status=200)

            # Serialize the formats
            from .serializers import EDAFormatSerializer
            serializer = EDAFormatSerializer(page_obj.object_list, many=True)
            
            return Response({
                'formats': serializer.data,
                'page': page,
                'total_pages': paginator.num_pages,
                'total_formats': paginator.count
            }, status=200)

        except Exception as e:
            return Response({'error': str(e)}, status=500)

class ApplyEDAFormat(APIView):
    """
    API to apply an EDA format to a dataset and create plots
    """
    parser_classes = [JSONParser]

    def post(self, request):
        try:
            # Extract data from request
            user_id = request.data.get('user_id')
            project_id = request.data.get('project_id')
            format_id = request.data.get('format_id')
            available_columns = request.data.get('available_columns', [])
            
            # Data source information
            file_type = request.data.get('file_type')
            file_name = request.data.get('file_name')
            sheet_name = request.data.get('sheet_name')

            # Validate required fields
            if not all([user_id, project_id, format_id]):
                missing_fields = []
                if not user_id: missing_fields.append('user_id')
                if not project_id: missing_fields.append('project_id')
                if not format_id: missing_fields.append('format_id')
                return Response({
                    'error': f'Missing required fields: {", ".join(missing_fields)}'
                }, status=400)

            # Get user, project, and format
            try:
                from .models import User
                user = User.objects.get(id=user_id)
                project = Projects.objects.get(id=project_id)
                eda_format = EDAFormat.objects.get(id=format_id)
            except User.DoesNotExist:
                return Response({'error': 'User not found'}, status=404)
            except Projects.DoesNotExist:
                return Response({'error': 'Project not found'}, status=404)
            except EDAFormat.DoesNotExist:
                return Response({'error': 'EDA format not found'}, status=404)

            # Check if user has access to this project
            if project.user != user:
                return Response({
                    'error': 'Access denied. You don\'t have permission to access this project.'
                }, status=403)

            # Check if user has access to this format
            if eda_format.user != user:
                return Response({
                    'error': 'Access denied. You don\'t have permission to use this format.'
                }, status=403)

            # Check column compatibility
            is_compatible, missing_columns, matched_columns = eda_format.check_column_compatibility(available_columns)
            
            if not is_compatible:
                return Response({
                    'error': 'Column compatibility check failed',
                    'missing_columns': missing_columns,
                    'available_columns': available_columns,
                    'required_columns': eda_format.required_columns
                }, status=400)

            # Apply the format and create plots
            created_plots = []
            format_config = eda_format.format_config
            
            print(f"Applying format '{eda_format.format_name}' with config: {format_config}")
            print(f"Available columns: {available_columns}")
            print(f"Column patterns: {eda_format.column_patterns}")
            
            if 'plots' in format_config:
                for plot_config in format_config['plots']:
                    try:
                        print(f"Processing plot config: {plot_config}")
                        
                        # Extract the actual plot configuration from the nested structure
                        actual_plot_config = plot_config.get('plot_config', plot_config)
                        print(f"Actual plot config: {actual_plot_config}")
                        
                        # Map column names from format to available columns
                        mapped_config = self._map_columns_to_available(actual_plot_config, available_columns, eda_format.column_patterns)
                        print(f"Mapped config: {mapped_config}")
                        
                        if mapped_config:
                            # Create the plot
                            plot_name = f"{eda_format.format_name} - {plot_config.get('name', 'Plot')}"
                            
                            # Handle Y-axes configuration for format-applied plots
                            x_axes = mapped_config.get('xAxes', [])
                            y_axes = mapped_config.get('yAxes', [])
                            y_axis = mapped_config.get('yAxis')
                            
                            # Ensure both formats are available for compatibility
                            if y_axes and not y_axis:
                                y_axis = y_axes[0] if y_axes else None
                            elif y_axis and not y_axes:
                                y_axes = [y_axis] if y_axis else []
                            
                            # Update mapped_config to ensure compatibility
                            mapped_config['yAxis'] = y_axis
                            mapped_config['yAxes'] = y_axes
                            
                            # Create columns_analyzed list
                            columns_analyzed = list(set(x_axes + y_axes))
                            
                            print(f"Creating plot '{plot_name}' with config:")
                            print(f"  - X-axes: {x_axes}")
                            print(f"  - Y-axes: {y_axes}")
                            print(f"  - Y-axis: {y_axis}")
                            print(f"  - Chart type: {mapped_config.get('chartType', 'bar')}")
                            print(f"  - Columns analyzed: {columns_analyzed}")
                            
                            eda_plot = EDAPlot.objects.create(
                                user=user,
                                project=project,
                                plot_name=plot_name,
                                plot_type=plot_config.get('plot_type', 'custom'),
                                file_type=file_type,
                                file_name=file_name,
                                sheet_name=sheet_name,
                                plot_config=mapped_config,
                                chart_data={},  # Will be populated when data is processed
                                chart_options=plot_config.get('chart_options', {}),
                                eda_analysis_type='format_applied',
                                columns_analyzed=columns_analyzed,
                                data_summary={
                                    'format_applied': eda_format.format_name,
                                    'format_id': eda_format.id,
                                    'applied_at': timezone.now().isoformat(),
                                    'x_axes': x_axes,
                                    'y_axis': y_axis,
                                    'y_axes': y_axes,
                                    'chart_type': mapped_config.get('chartType', 'bar'),
                                    'aggregation_method': mapped_config.get('aggregationMethod', 'sum'),
                                    'date_grouping': mapped_config.get('dateGrouping', 'raw')
                                },
                                description=f"Auto-generated from format: {eda_format.format_name}",
                                tags=eda_format.tags + ['auto-generated', 'format-applied']
                            )
                            
                            print(f"Successfully created plot with ID: {eda_plot.id}")
                            
                            created_plots.append({
                                'plot_id': eda_plot.id,
                                'plot_name': eda_plot.plot_name,
                                'plot_config': mapped_config
                            })
                        else:
                            print(f"Failed to map plot configuration for plot: {plot_config.get('name', 'Unknown')}")
                    except Exception as e:
                        # Continue with other plots even if one fails
                        print(f"Error creating plot from format: {e}")

            # Increment format usage
            eda_format.increment_usage()

            # Log the action
            ip = request.META.get('REMOTE_ADDR')
            from .log_utils import log_user_action
            log_user_action(user, "apply_eda_format", details=f"Applied format '{eda_format.format_name}' to project {project.name}", ip_address=ip)

            return Response({
                'message': f'Successfully applied format and created {len(created_plots)} plots',
                'format_name': eda_format.format_name,
                'created_plots': created_plots,
                'matched_columns': matched_columns,
                'usage_count': eda_format.usage_count
            }, status=200)

        except Exception as e:
            return Response({'error': str(e)}, status=500)

    def _map_columns_to_available(self, plot_config, available_columns, column_patterns):
        """
        Map column names from format to available columns using patterns
        """
        print(f"Mapping plot config: {plot_config}")
        print(f"Available columns: {available_columns}")
        print(f"Column patterns: {column_patterns}")
        
        mapped_config = plot_config.copy()
        
        # Map X-axes
        if 'xAxes' in plot_config:
            mapped_x_axes = []
            for x_axis in plot_config['xAxes']:
                mapped_col = self._find_matching_column(x_axis, available_columns, column_patterns)
                print(f"Mapping X-axis '{x_axis}' to '{mapped_col}'")
                if mapped_col:
                    mapped_x_axes.append(mapped_col)
                else:
                    print(f"Failed to map X-axis '{x_axis}'")
                    return None  # Cannot map this plot
            mapped_config['xAxes'] = mapped_x_axes
        
        # Map Y-axes (handle both yAxis and yAxes)
        if 'yAxes' in plot_config:
            mapped_y_axes = []
            for y_axis in plot_config['yAxes']:
                mapped_col = self._find_matching_column(y_axis, available_columns, column_patterns)
                print(f"Mapping Y-axis '{y_axis}' to '{mapped_col}'")
                if mapped_col:
                    mapped_y_axes.append(mapped_col)
                else:
                    print(f"Failed to map Y-axis '{y_axis}'")
                    return None  # Cannot map this plot
            mapped_config['yAxes'] = mapped_y_axes
            
            # Also set yAxis for backward compatibility
            if mapped_y_axes:
                mapped_config['yAxis'] = mapped_y_axes[0]
        
        # Map Y-axis (fallback for backward compatibility)
        elif 'yAxis' in plot_config:
            mapped_y_axis = self._find_matching_column(plot_config['yAxis'], available_columns, column_patterns)
            print(f"Mapping Y-axis '{plot_config['yAxis']}' to '{mapped_y_axis}'")
            if mapped_y_axis:
                mapped_config['yAxis'] = mapped_y_axis
                # Also set yAxes for consistency
                mapped_config['yAxes'] = [mapped_y_axis]
            else:
                print(f"Failed to map Y-axis '{plot_config['yAxis']}'")
                return None  # Cannot map this plot
        
        print(f"Final mapped config: {mapped_config}")
        return mapped_config

    def _find_matching_column(self, target_column, available_columns, column_patterns):
        """
        Find a matching column using exact match or patterns
        """
        print(f"Finding match for column '{target_column}' in available columns: {available_columns}")
        
        # First try exact match
        if target_column in available_columns:
            print(f"Exact match found: '{target_column}'")
            return target_column
        
        # Then try patterns
        if target_column in column_patterns:
            pattern_config = column_patterns[target_column]
            pattern_type = pattern_config.get('type', 'exact')
            pattern_value = pattern_config.get('value', '')
            
            print(f"Trying pattern match - type: {pattern_type}, value: '{pattern_value}'")
            
            if pattern_type == 'contains':
                for col in available_columns:
                    if pattern_value.lower() in col.lower():
                        print(f"Pattern match found: '{col}' contains '{pattern_value}'")
                        return col
            elif pattern_type == 'regex':
                import re
                regex = re.compile(pattern_value, re.IGNORECASE)
                for col in available_columns:
                    if regex.search(col):
                        print(f"Regex match found: '{col}' matches '{pattern_value}'")
                        return col
        
        print(f"No match found for column '{target_column}'")
        return None


class DeleteEDAFormat(APIView):
    """
    API to delete EDA formats
    """
    parser_classes = [JSONParser]

    def delete(self, request, format_id):
        try:
            # Extract data from query parameters (DELETE requests don't have body)
            user_id = request.query_params.get('user_id')
            
            if not user_id:
                return Response({
                    'error': 'Missing required parameter: user_id'
                }, status=400)
            
            # Get the format
            try:
                eda_format = EDAFormat.objects.get(id=format_id)
            except EDAFormat.DoesNotExist:
                return Response({'error': 'EDA format not found'}, status=404)

            # Check if user has permission to delete this format
            if eda_format.user.id != int(user_id):
                return Response({
                    'error': 'Access denied. You don\'t have permission to delete this format.'
                }, status=403)

            format_name = eda_format.format_name
            eda_format.delete()

            # Log the action
            ip = request.META.get('REMOTE_ADDR')
            from .log_utils import log_user_action
            from .models import User
            user = User.objects.get(id=user_id)
            log_user_action(user, "delete_eda_format", details=f"EDA format '{format_name}' deleted", ip_address=ip)

            return Response({
                'message': 'EDA format deleted successfully'
            }, status=200)

        except Exception as e:
            return Response({'error': str(e)}, status=500)

class GetEDAFormatDetails(APIView):
    """
    API to get detailed information about a specific EDA format
    """
    parser_classes = [JSONParser]

    def post(self, request):
        try:
            # Extract data from request
            user_id = request.data.get('user_id')
            format_id = request.data.get('format_id')
            
            if not all([user_id, format_id]):
                missing_fields = []
                if not user_id: missing_fields.append('user_id')
                if not format_id: missing_fields.append('format_id')
                return Response({
                    'error': f'Missing required fields: {", ".join(missing_fields)}'
                }, status=400)

            # Get the format
            try:
                eda_format = EDAFormat.objects.get(id=format_id)
            except EDAFormat.DoesNotExist:
                return Response({'error': 'EDA format not found'}, status=404)

            # Check if user has permission to view this format
            # Convert user_id to integer for comparison
            try:
                user_id_int = int(user_id)
            except (ValueError, TypeError):
                return Response({
                    'error': 'Invalid user_id format'
                }, status=400)
            
            if eda_format.user.id != user_id_int:
                return Response({
                    'error': 'Access denied. You don\'t have permission to view this format.'
                }, status=403)

            from .serializers import EDAFormatSerializer
            serializer = EDAFormatSerializer(eda_format)
            return Response({
                'format': serializer.data
            }, status=200)

        except Exception as e:
            return Response({'error': str(e)}, status=500) 
    
class DownloadEDAPlotsExcel(APIView):
    """
    API to download all saved EDA plots with their data in an Excel file
    The Excel file will contain editable charts that can be modified in Microsoft Excel
    """
    parser_classes = [JSONParser]

    def post(self, request):
        try:
            # Extract data from request
            user_id = request.data.get('user_id')
            project_id = request.data.get('project_id')
            plot_ids = request.data.get('plot_ids', [])  # Optional: specific plot IDs to download
            # Note: Empty array means download all plots (this is intentional behavior)
            
            # Validate required fields
            if not all([user_id, project_id]):
                missing_fields = []
                if not user_id: missing_fields.append('user_id')
                if not project_id: missing_fields.append('project_id')
                return Response({
                    'error': f'Missing required fields: {", ".join(missing_fields)}'
                }, status=400)

            # Get user and project
            try:
                from .models import User
                user = User.objects.get(id=user_id)
                project = Projects.objects.get(id=project_id)
            except User.DoesNotExist:
                return Response({'error': 'User not found'}, status=404)
            except Projects.DoesNotExist:
                return Response({'error': 'Project not found'}, status=404)

            # Check if user has access to this project
            if project.user != user:
                return Response({
                    'error': 'Access denied. You don\'t have permission to access this project.'
                }, status=403)

            # Get EDA plots
            plots_query = EDAPlot.objects.filter(user=user, project=project)
            
            if plot_ids:
                # If specific plot IDs are provided, filter by them
                plots_query = plots_query.filter(id__in=plot_ids)
            # If plot_ids is empty or not provided, get all plots (this is the intended behavior)
            
            eda_plots = plots_query.order_by('created_at')
            
            if not eda_plots.exists():
                return Response({
                    'error': 'No plots found for this project.'
                }, status=404)

            # Create Excel workbook
            import pandas as pd
            from django.http import HttpResponse
            import json

            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create summary sheet first
            ws_summary = wb.create_sheet("Summary")
            ws_summary.append(['EDA Plots Summary'])
            ws_summary.append([''])
            ws_summary.append(['Plot Name', 'Chart Type', 'X-Axes', 'Y-Axes', 'Created Date', 'Description'])
            
            # Style the header
            for cell in ws_summary[3]:
                cell.font = cell.font.copy(bold=True)
                cell.fill = cell.fill.copy(fill_type="solid", fgColor="CCCCCC")

            plot_count = 0
            
            for plot in eda_plots:
                plot_count += 1
                
                # Add to summary
                x_axes = ', '.join(plot.get_x_axes()) if plot.get_x_axes() else 'N/A'
                y_axes = ', '.join(plot.get_y_axes()) if plot.get_y_axes() else 'N/A'
                chart_type = plot.get_chart_type()
                # Add special indicator for stacked charts
                if chart_type in ['stacked', 'stacked100']:
                    chart_type_display = f"{chart_type.upper()} ({len(plot.get_y_axes())} series)"
                else:
                    chart_type_display = chart_type
                ws_summary.append([
                    plot.plot_name,
                    chart_type_display,
                    x_axes,
                    y_axes,
                    plot.created_at.strftime('%Y-%m-%d %H:%M'),
                    plot.description or 'No description'
                ])
                
                # Create sheet for this plot
                sheet_name = f"Plot_{plot_count}_{plot.plot_name[:20]}"  # Excel sheet name limit
                sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '_', '-')).rstrip()
                if len(sheet_name) > 31:  # Excel sheet name length limit
                    sheet_name = sheet_name[:31]
                
                ws_plot = wb.create_sheet(sheet_name)
                
                # Get plot configuration
                plot_config = plot.plot_config
                chart_data = plot.chart_data
                
                # Extract axes information
                x_axes = plot.get_x_axes()
                y_axes = plot.get_y_axes()
                chart_type = plot.get_chart_type()
                aggregation_method = plot.get_aggregation_method()
                date_grouping = plot.get_date_grouping()
                
                # Process data for Excel - similar to EDAPlotSheetData.jsx logic
                processed_data = self._process_plot_data_for_excel(
                    plot, x_axes, y_axes, chart_type, aggregation_method, date_grouping
                )
                
                print(f"Processing plot '{plot.plot_name}':")
                print(f"  - X-axes: {x_axes}")
                print(f"  - Y-axes: {y_axes}")
                print(f"  - Chart type: {chart_type}")
                print(f"  - Aggregation method: {aggregation_method}")
                print(f"  - Processed data rows: {len(processed_data) if processed_data else 0}")
                
                if processed_data and len(processed_data) > 0:
                    # Pre-clean the data to remove any Period objects before DataFrame conversion
                    cleaned_processed_data = []
                    for row in processed_data:
                        cleaned_row = {}
                        for key, value in row.items():
                            # Clean the key (column name)
                            clean_key = str(key).replace('Period_', '')
                            # Clean the value
                            clean_value = self._safe_convert_value(value)
                            cleaned_row[clean_key] = clean_value
                        cleaned_processed_data.append(cleaned_row)
                    
                    # Convert to DataFrame
                    df_data = pd.DataFrame(cleaned_processed_data)
                    
                    # Debug: Print data types before cleaning
                    print(f"    - DataFrame columns and dtypes:")
                    for col in df_data.columns:
                        print(f"      {col}: {df_data[col].dtype}")
                    
                    # Debug: Print sample values to identify Period objects
                    print(f"    - Sample values from first few rows:")
                    for col in df_data.columns:
                        sample_vals = df_data[col].dropna().head(3).tolist()
                        print(f"      {col}: {sample_vals}")
                    
                    # Additional cleaning for DataFrame
                    df_data = self._clean_dataframe_for_excel(df_data)
                    
                    # Clean the X-axes and Y-axes arrays to match the cleaned DataFrame columns
                    cleaned_x_axes = [str(axis).replace('Period_', '') for axis in x_axes]
                    cleaned_y_axes = [str(axis).replace('Period_', '') for axis in y_axes]
                    
                    # Convert Y-axis columns to numeric for proper chart display
                    for col in df_data.columns:
                        if col in cleaned_y_axes:
                            try:
                                df_data[col] = pd.to_numeric(df_data[col], errors='coerce')
                                print(f"Converted column '{col}' to numeric, sample values: {df_data[col].head().tolist()}")
                            except Exception as conv_error:
                                print(f"Error converting column '{col}' to numeric: {conv_error}")
                    
                    # Debug: Print data types after cleaning
                    print(f"    - DataFrame columns and dtypes after cleaning:")
                    for col in df_data.columns:
                        print(f"      {col}: {df_data[col].dtype}")
                    
                    # Write data to Excel sheet
                    try:
                        for r in dataframe_to_rows(df_data, index=False, header=True):
                            ws_plot.append(r)
                        
                        # Debug: Print the actual data written to Excel
                        print(f"Data written to Excel sheet '{sheet_name}':")
                        for row_idx, row in enumerate(ws_plot.iter_rows(min_row=1, max_row=min(5, len(df_data)+1)), 1):
                            row_data = [cell.value for cell in row]
                            print(f"  Row {row_idx}: {row_data}")
                        
                    except Exception as excel_error:
                        print(f"Error writing DataFrame to Excel for plot {plot.plot_name}: {excel_error}")
                        print(f"DataFrame info: {df_data.info()}")
                        print(f"DataFrame head: {df_data.head()}")
                        
                        # Try with a simple fallback approach
                        print(f"Trying fallback approach...")
                        try:
                            # Create a simple DataFrame with just strings
                            simple_data = []
                            for row in cleaned_processed_data:
                                simple_row = {}
                                for key, value in row.items():
                                    simple_row[str(key)] = str(value)
                                simple_data.append(simple_row)
                            
                            df_simple = pd.DataFrame(simple_data)
                            print(f"Simple DataFrame created with {len(df_simple)} rows and {len(df_simple.columns)} columns")
                            
                            for r in dataframe_to_rows(df_simple, index=False, header=True):
                                ws_plot.append(r)
                            print(f"Simple DataFrame written successfully")
                            
                        except Exception as fallback_error:
                            print(f"Fallback also failed: {fallback_error}")
                            # Ultimate fallback: create a basic sheet with error message
                            ws_plot.append(['Error: Could not process data'])
                            ws_plot.append(['Data contained incompatible types'])
                            ws_plot.append(['Please check the original data source'])
                            
                            # Try with completely clean test data
                            print(f"Trying with clean test data...")
                            try:
                                df_test = self._create_clean_test_dataframe(plot.plot_name)
                                for r in dataframe_to_rows(df_test, index=False, header=True):
                                    ws_plot.append(r)
                                print(f"Test data written successfully")
                            except Exception as test_error:
                                print(f"Test data also failed: {test_error}")
                                ws_plot.append(['All data processing methods failed'])
                    
                    # Style the header
                    for cell in ws_plot[1]:
                        cell.font = cell.font.copy(bold=True)
                        cell.fill = cell.fill.copy(fill_type="solid", fgColor="E6E6E6")
                    
                    # Create chart based on chart type
                    chart = self._create_excel_chart(chart_type, plot.plot_name)
                    
                    if chart and len(df_data) > 0 and len(df_data.columns) > 0:
                        # Configure chart data
                        max_row = len(df_data) + 1  # +1 for header
                        max_col = len(df_data.columns)
                        
                        # Find column indices for X and Y axes
                        x_col_indices = []
                        y_col_indices = []
                        
                        print(f"Original X-axes: {x_axes}, Cleaned X-axes: {cleaned_x_axes}")
                        print(f"Original Y-axes: {y_axes}, Cleaned Y-axes: {cleaned_y_axes}")
                        print(f"Available DataFrame columns: {list(df_data.columns)}")
                        
                        for i, col in enumerate(df_data.columns, 1):
                            if col in cleaned_x_axes:
                                x_col_indices.append(i)
                            if col in cleaned_y_axes:
                                y_col_indices.append(i)
                        
                        print(f"Found X column indices: {x_col_indices}")
                        print(f"Found Y column indices: {y_col_indices}")
                        
                        if x_col_indices and y_col_indices:
                            try:
                                # For multiple X-axes, use the first one for categories
                                cats = Reference(ws_plot, min_col=x_col_indices[0], min_row=2, max_row=max_row)
                                chart.set_categories(cats)
                                
                                # Enhanced data addition for stacked charts
                                if chart_type in ['stacked', 'stacked100']:
                                    print(f"    - Adding data for stacked chart: {chart_type}")
                                    # For stacked charts, add each Y-axis as a separate series
                                    for i, y_col_idx in enumerate(y_col_indices):
                                        data = Reference(ws_plot, min_col=y_col_idx, min_row=1, max_row=max_row)
                                        chart.add_data(data, titles_from_data=True)
                                        print(f"      - Added series {i+1}: column {y_col_idx} ({cleaned_y_axes[i] if i < len(cleaned_y_axes) else 'Unknown'})")
                                else:
                                    # For regular charts, add data as before
                                    for y_col_idx in y_col_indices:
                                        data = Reference(ws_plot, min_col=y_col_idx, min_row=1, max_row=max_row)
                                        chart.add_data(data, titles_from_data=True)
                                
                                # Add chart to sheet with specific positioning and size
                                # Use a more reliable method to set chart size
                                try:
                                    # Create anchor with specific size (EMU = Excel's internal units)
                                    # 1 inch = 914400 EMUs
                                    width_emu = int(9.2 * 914400)  # 9.2 inches in EMUs
                                    height_emu = int(4.2 * 914400)  # 4.2 inches in EMUs
                                    
                                    anchor = OneCellAnchor(
                                        _from=AnchorMarker(col=4, colOff=0, row=1, rowOff=0),  # Position at E2
                                        ext={'cx': width_emu, 'cy': height_emu}  # Width=9.2", Height=4.2"
                                    )
                                    ws_plot.add_chart(chart, anchor)
                                    print(f"Chart added with explicit size: width=9.2\", height=4.2\"")
                                except Exception as anchor_error:
                                    print(f"Error with anchor method: {anchor_error}")
                                    # Fallback to simple add_chart method
                                    ws_plot.add_chart(chart, f"E2")
                                    print(f"Chart added with default positioning")
                                    
                                    # Try to set size after adding
                                    try:
                                        # Get the chart object from the worksheet
                                        for chart_obj in ws_plot._charts:
                                            if hasattr(chart_obj, 'graphicFrame'):
                                                if hasattr(chart_obj.graphicFrame, 'xfrm'):
                                                    if hasattr(chart_obj.graphicFrame.xfrm, 'ext'):
                                                        # Set chart size: width=9.2, height=4.2 (in inches)
                                                        chart_obj.graphicFrame.xfrm.ext.cx = int(9.2 * 914400)
                                                        chart_obj.graphicFrame.xfrm.ext.cy = int(4.2 * 914400)
                                                        print(f"Chart size set successfully: width=9.2\", height=4.2\"")
                                                        break
                                    except Exception as size_error:
                                        print(f"Error setting chart size: {size_error}")
                                        # Final fallback: try to set size on the original chart object
                                        try:
                                            if hasattr(chart, 'graphicFrame'):
                                                if hasattr(chart.graphicFrame, 'xfrm'):
                                                    if hasattr(chart.graphicFrame.xfrm, 'ext'):
                                                        chart.graphicFrame.xfrm.ext.cx = int(9.2 * 914400)
                                                        chart.graphicFrame.xfrm.ext.cy = int(4.2 * 914400)
                                                        print(f"Chart size set via final fallback method")
                                        except Exception as fallback_error:
                                            print(f"All chart size setting methods failed: {fallback_error}")
                                
                                print(f"Chart created successfully for plot {plot.plot_name}")
                                print(f"Chart data range: X={x_col_indices[0]}:2-{max_row}, Y={y_col_indices}:1-{max_row}")
                                print(f"Chart size set to: width=9.2\", height=4.2\"")
                                print(f"Chart type: {chart_type}")
                            except Exception as chart_error:
                                print(f"Error creating chart for plot {plot.plot_name}: {chart_error}")
                                print(f"Chart type: {chart_type}, X columns: {x_col_indices}, Y columns: {y_col_indices}")
                                print(f"DataFrame shape: {df_data.shape}")
                                print(f"DataFrame head: {df_data.head()}")
                        else:
                            print(f"No valid X or Y columns found for chart in plot {plot.plot_name}")
                            print(f"X columns needed: {cleaned_x_axes}, Y columns needed: {cleaned_y_axes}")
                            print(f"Available columns: {list(df_data.columns)}")
                    else:
                        print(f"Chart creation skipped for plot {plot.plot_name} - no chart object or no data")
                        # Initialize cleaned axes for metadata even if no chart
                        cleaned_x_axes = [str(axis).replace('Period_', '') for axis in x_axes]
                        cleaned_y_axes = [str(axis).replace('Period_', '') for axis in y_axes]
                
                # Add plot metadata
                ws_plot.append([])
                ws_plot.append(['Plot Metadata'])
                ws_plot.append(['Property', 'Value'])
                ws_plot.append(['Plot Name', plot.plot_name])
                ws_plot.append(['Chart Type', chart_type])
                ws_plot.append(['X-Axes', ', '.join(cleaned_x_axes)])
                ws_plot.append(['Y-Axes', ', '.join(cleaned_y_axes)])
                ws_plot.append(['Aggregation Method', aggregation_method])
                ws_plot.append(['Date Grouping', date_grouping])
                ws_plot.append(['Created Date', plot.created_at.strftime('%Y-%m-%d %H:%M:%S')])
                ws_plot.append(['Updated Date', plot.updated_at.strftime('%Y-%m-%d %H:%M:%S')])
                if plot.description:
                    ws_plot.append(['Description', plot.description])
                if plot.tags:
                    ws_plot.append(['Tags', ', '.join(plot.tags)])
                
                # Style metadata section
                metadata_start_row = max_row + 3 if 'max_row' in locals() else 15
                for row in range(metadata_start_row, metadata_start_row + 10):
                    for col in range(1, 3):
                        cell = ws_plot.cell(row=row, column=col)
                        if col == 1:  # Property column
                            cell.font = cell.font.copy(bold=True)
                            cell.fill = cell.fill.copy(fill_type="solid", fgColor="F0F0F0")


            
            # Create instructions sheet after plot count is determined
            ws_instructions = wb.create_sheet("Instructions")
            ws_instructions.append(['EDA Plots Excel File - Instructions'])
            ws_instructions.append([''])
            ws_instructions.append(['This Excel file contains all your saved EDA plots with their data.'])
            ws_instructions.append([''])
            ws_instructions.append(['How to use:'])
            ws_instructions.append(['1. Each plot has its own sheet with the data used to create the chart'])
            ws_instructions.append(['2. The charts are editable - you can modify the data and the charts will update automatically'])
            ws_instructions.append(['3. To edit chart data:'])
            ws_instructions.append(['   - Click on the chart to select it'])
            ws_instructions.append(['   - Right-click and select "Edit Data" or "Select Data"'])
            ws_instructions.append(['   - Modify the data range or values as needed'])
            ws_instructions.append(['4. To modify chart appearance:'])
            ws_instructions.append(['   - Right-click on the chart and select "Format Chart Area"'])
            ws_instructions.append(['   - Use the Chart Tools ribbon for additional formatting options'])
            ws_instructions.append(['5. Stacked Charts:'])
            ws_instructions.append(['   - Stacked charts show multiple data series stacked on top of each other'])
            ws_instructions.append(['   - Stacked 100% charts show the percentage contribution of each series'])
            ws_instructions.append(['   - Each Y-axis column represents a different series in the stack'])
            ws_instructions.append([''])
            ws_instructions.append(['File Information:'])
            ws_instructions.append(['Project:', project.name])
            ws_instructions.append(['Total Plots:', str(plot_count)])
            ws_instructions.append(['Generated:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
            
            # Style instructions
            for row in range(1, 20):
                for col in range(1, 2):
                    cell = ws_instructions.cell(row=row, column=col)
                    if row == 1:
                        cell.font = cell.font.copy(bold=True, size=14)
                    elif row in [3, 5, 12, 16]:
                        cell.font = cell.font.copy(bold=True)
            
            # Reorder sheets: Instructions first, then Summary, then plot sheets
            # Get all sheet names in current order
            sheet_names = wb.sheetnames
            
            # Create the desired order: Instructions, Summary, then plot sheets
            desired_order = ['Instructions', 'Summary']
            
            # Add all plot sheets to the desired order (they should already be in the correct order)
            for sheet_name in sheet_names:
                if sheet_name not in desired_order:
                    desired_order.append(sheet_name)
            
            # Reorder sheets by moving them to the desired positions
            # First, move Instructions to position 0 (first sheet)
            if 'Instructions' in wb.sheetnames:
                wb.move_sheet('Instructions', offset=-wb.sheetnames.index('Instructions'))
            
            # Then, move Summary to position 1 (second sheet)
            if 'Summary' in wb.sheetnames:
                current_summary_index = wb.sheetnames.index('Summary')
                if current_summary_index != 1:  # Only move if not already in position 1
                    wb.move_sheet('Summary', offset=1-current_summary_index)
            
            # Save to BytesIO instead of temporary file to avoid file locking issues
            excel_file = BytesIO()
            try:
                wb.save(excel_file)
                excel_file.seek(0)
                
                # Get the file content
                file_content = excel_file.getvalue()
                
                # Create response
                response = HttpResponse(
                    file_content,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="eda_plots_{project.name}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
                response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
                response['Pragma'] = 'no-cache'
                response['Expires'] = '0'
                
                # Log the action
                ip = request.META.get('REMOTE_ADDR')
                from .log_utils import log_user_action
                log_user_action(user, "download_eda_plots_excel", 
                              details=f"Downloaded {plot_count} EDA plots as Excel file for project {project.name}", 
                              ip_address=ip)
                
                return response
                
            finally:
                # Always close the BytesIO object
                excel_file.close()

        except Exception as e:
            return Response({'error': str(e)}, status=500)

    def _process_plot_data_for_excel(self, plot, x_axes, y_axes, chart_type, aggregation_method, date_grouping):
        """
        Process plot data similar to EDAPlotSheetData.jsx logic with enhanced stacked chart support
        """
        try:
            print(f"  _process_plot_data_for_excel called for plot '{plot.plot_name}'")
            print(f"    - Chart type: {chart_type}")
            print(f"    - X-axes: {x_axes}")
            print(f"    - Y-axes: {y_axes}")
            
            # Get the original sheet data from the plot's data_summary or try to reconstruct
            chart_data = plot.chart_data
            plot_config = plot.plot_config
            
            print(f"    - Has chart_data: {bool(chart_data)}")
            print(f"    - Chart data length: {len(chart_data) if chart_data else 0}")
            
            # If we have processed chart_data, use it
            if chart_data and isinstance(chart_data, list) and len(chart_data) > 0:
                print(f"    - Using existing chart_data")
                # Clean and validate the data
                processed_data = []
                for row in chart_data:
                    if isinstance(row, dict):
                        # Remove internal fields and clean data
                        clean_row = {}
                        for key, value in row.items():
                            if not key.startswith('__'):
                                # Clean the key first
                                clean_key = str(key).replace('Period_', '')
                                # Handle Period objects and other non-serializable types
                                if hasattr(value, 'to_timestamp'):  # Period objects
                                    clean_row[clean_key] = value.to_timestamp().strftime('%Y-%m-%d')
                                elif hasattr(value, 'isoformat'):  # Timestamp objects
                                    clean_row[clean_key] = value.isoformat()
                                elif isinstance(value, (list, dict)):
                                    clean_row[clean_key] = str(value)
                                elif isinstance(value, str) and value.startswith('Period_'):
                                    clean_row[clean_key] = value.replace('Period_', '')
                                else:
                                    clean_row[clean_key] = value
                        processed_data.append(clean_row)
                
                # Clean any Period objects in the data
                for row in processed_data:
                    for key, value in row.items():
                        if isinstance(value, str) and value.startswith('Period_'):
                            row[key] = value.replace('Period_', '')
                        elif hasattr(value, 'strftime'):  # Handle datetime objects
                            row[key] = value.strftime('%Y-%m-%d %H:%M:%S')
                
                print(f"    - Cleaned chart_data rows: {len(processed_data)}")
                
                # For stacked charts, ensure data is properly formatted
                if chart_type in ['stacked', 'stacked100'] and processed_data:
                    processed_data = self._format_data_for_stacked_chart(processed_data, x_axes, y_axes, chart_type)
                
                return processed_data
            
            # Try to get original data from the project file
            print(f"    - Attempting to get original sheet data")
            original_data = self._get_original_sheet_data(plot)
            if original_data and len(original_data) > 0:
                print(f"    - Got original data, processing for chart")
                # Process the original data similar to EDAPlotSheetData.jsx
                processed_data = self._process_raw_data_for_chart(
                    original_data, x_axes, y_axes, aggregation_method, date_grouping
                )
                
                # Clean any Period objects in the processed data
                if processed_data:
                    for row in processed_data:
                        for key, value in row.items():
                            if isinstance(value, str) and value.startswith('Period_'):
                                row[key] = value.replace('Period_', '')
                            elif hasattr(value, 'strftime'):
                                row[key] = value.strftime('%Y-%m-%d %H:%M:%S')
                
                # For stacked charts, ensure data is properly formatted
                if chart_type in ['stacked', 'stacked100'] and processed_data:
                    processed_data = self._format_data_for_stacked_chart(processed_data, x_axes, y_axes, chart_type)
                
                return processed_data
            
            # If no chart_data, try to get from data_summary
            data_summary = plot.data_summary or {}
            if 'raw_data' in data_summary:
                print(f"    - Using raw_data from data_summary")
                processed_data = data_summary['raw_data']
                # Clean any Period objects in the raw data
                if processed_data:
                    for row in processed_data:
                        for key, value in row.items():
                            if isinstance(value, str) and value.startswith('Period_'):
                                row[key] = value.replace('Period_', '')
                            elif hasattr(value, 'strftime'):
                                row[key] = value.strftime('%Y-%m-%d %H:%M:%S')
                
                # For stacked charts, ensure data is properly formatted
                if chart_type in ['stacked', 'stacked100'] and processed_data:
                    processed_data = self._format_data_for_stacked_chart(processed_data, x_axes, y_axes, chart_type)
                
                return processed_data
            
            # Fallback: create sample data based on plot configuration
            if x_axes and y_axes:
                print(f"    - Creating sample data")
                sample_data = []
                for i in range(10):  # Create 10 sample data points
                    row = {}
                    for x_axis in x_axes:
                        row[x_axis] = f"Sample {i+1}"
                    for y_axis in y_axes:
                        row[y_axis] = (i + 1) * 10  # Sample values
                    sample_data.append(row)
                
                # For stacked charts, ensure sample data is properly formatted
                if chart_type in ['stacked', 'stacked100']:
                    sample_data = self._format_data_for_stacked_chart(sample_data, x_axes, y_axes, chart_type)
                
                return sample_data
            
            print(f"    - No data available")
            return []
            
        except Exception as e:
            print(f"Error processing data for plot {plot.plot_name}: {e}")
            return []

    def _format_data_for_stacked_chart(self, data, x_axes, y_axes, chart_type):
        """
        Format data specifically for stacked charts to ensure proper Excel chart rendering
        """
        try:
            print(f"    - Formatting data for stacked chart type: {chart_type}")
            
            if not data or not x_axes or not y_axes:
                return data
            
            # For stacked charts, we need to ensure:
            # 1. All Y-axis values are numeric
            # 2. For stacked100, values should be percentages
            # 3. Data structure is consistent
            
            formatted_data = []
            
            for row in data:
                formatted_row = {}
                
                # Copy X-axis values as-is
                for x_axis in x_axes:
                    if x_axis in row:
                        formatted_row[x_axis] = row[x_axis]
                    else:
                        formatted_row[x_axis] = ''
                
                # Process Y-axis values
                y_values = []
                for y_axis in y_axes:
                    if y_axis in row:
                        try:
                            # Convert to numeric value
                            val = float(row[y_axis]) if row[y_axis] is not None else 0.0
                            y_values.append(val)
                        except (ValueError, TypeError):
                            y_values.append(0.0)
                    else:
                        y_values.append(0.0)
                
                # For stacked100, normalize values to percentages
                if chart_type == 'stacked100' and y_values:
                    total = sum(y_values)
                    if total > 0:
                        y_values = [(val / total) * 100 for val in y_values]
                    else:
                        y_values = [0.0] * len(y_values)
                
                # Add Y-axis values to formatted row
                for i, y_axis in enumerate(y_axes):
                    formatted_row[y_axis] = y_values[i]
                
                formatted_data.append(formatted_row)
            
            print(f"    - Formatted {len(formatted_data)} rows for stacked chart")
            return formatted_data
            
        except Exception as e:
            print(f"Error formatting data for stacked chart: {e}")
            return data

    def _get_original_sheet_data(self, plot):
        """
        Try to get original sheet data from the project file
        """
        try:
            # Get file information from the plot
            file_type = plot.file_type
            file_name = plot.file_name
            sheet_name = plot.sheet_name
            project = plot.project
            
            if not all([file_type, file_name, sheet_name]):
                return None
            
            # Construct file path
            project_folder = f"user_{project.user.id}/project_{project.id}"
            
            if file_type == 'kpi':
                file_path = os.path.join(settings.MEDIA_ROOT, project_folder, "kpi", file_name)
            elif file_type == 'media':
                file_path = os.path.join(settings.MEDIA_ROOT, project_folder, "media", file_name)
            else:
                file_path = os.path.join(settings.MEDIA_ROOT, project_folder, file_name)
            
            if not os.path.exists(file_path):
                return None
            
            # Read the file
            file_extension = os.path.splitext(file_name)[1].lower()
            
            if file_extension == '.xlsx':
                df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
            elif file_extension == '.csv':
                try:
                    df = pd.read_csv(file_path, encoding='utf-8')
                except UnicodeDecodeError:
                    try:
                        df = pd.read_csv(file_path, encoding='latin1')
                    except UnicodeDecodeError:
                        df = pd.read_csv(file_path, encoding='cp1252')
            else:
                return None
            
            # Aggressively clean the DataFrame to handle Period objects
            df = self._aggressive_clean_dataframe(df)
            
            # Convert to list of dictionaries
            data = df.to_dict('records')
            return data
            
        except Exception as e:
            print(f"Error getting original sheet data for plot {plot.plot_name}: {e}")
            return None

    def _process_raw_data_for_chart(self, raw_data, x_axes, y_axes, aggregation_method, date_grouping):
        """
        Process raw data similar to EDAPlotSheetData.jsx logic
        """
        try:
            if not raw_data or not x_axes or not y_axes:
                return []
            
            # Get columns from first row
            columns = list(raw_data[0].keys()) if raw_data else []
            
            # Filter data based on available columns
            filtered_data = []
            for row in raw_data:
                if all(col in row for col in x_axes + y_axes):
                    filtered_data.append(row)
            
            if not filtered_data:
                return []
            
            # Aggregate data
            aggregation = {}
            for row in filtered_data:
                x_key = " | ".join(str(row.get(x, '')) for x in x_axes)
                
                if x_key not in aggregation:
                    aggregation[x_key] = {
                        'x_values': [row.get(x, '') for x in x_axes],
                        'y_values': {},
                        'count': 0
                    }
                
                # Process each Y-axis
                for y_axis in y_axes:
                    y_val = row.get(y_axis, 0)
                    try:
                        y_val = float(y_val) if y_val is not None else 0
                    except (ValueError, TypeError):
                        y_val = 0
                    
                    if y_axis not in aggregation[x_key]['y_values']:
                        aggregation[x_key]['y_values'][y_axis] = {
                            'sum': 0,
                            'min': y_val,
                            'max': y_val,
                            'count': 0
                        }
                    
                    agg = aggregation[x_key]['y_values'][y_axis]
                    agg['sum'] += y_val
                    agg['min'] = min(agg['min'], y_val)
                    agg['max'] = max(agg['max'], y_val)
                    agg['count'] += 1
                
                aggregation[x_key]['count'] += 1
            
            # Create final data
            processed_data = []
            for x_key, agg in aggregation.items():
                row = {}
                
                # Add X-axis values
                for i, x_axis in enumerate(x_axes):
                    row[x_axis] = agg['x_values'][i] if i < len(agg['x_values']) else ''
                
                # Add Y-axis values with aggregation
                for y_axis in y_axes:
                    if y_axis in agg['y_values']:
                        y_agg = agg['y_values'][y_axis]
                        if aggregation_method == 'avg':
                            row[y_axis] = y_agg['sum'] / y_agg['count'] if y_agg['count'] > 0 else 0
                        elif aggregation_method == 'max':
                            row[y_axis] = y_agg['max']
                        elif aggregation_method == 'min':
                            row[y_axis] = y_agg['min']
                        else:  # default to sum
                            row[y_axis] = y_agg['sum']
                    else:
                        row[y_axis] = 0
                
                processed_data.append(row)
            
            return processed_data
            
        except Exception as e:
            print(f"Error processing raw data: {e}")
            return []

    def _create_excel_chart(self, chart_type, plot_name):
        """
        Create an Excel chart with improved appearance and enhanced stacked chart support.
        - Axes, axis titles, chart title, legends, and data labels enabled by default.
        - Data table, error bars, gridlines, and trend lines disabled by default.
        - Enhanced support for stacked charts with proper grouping and percentage calculations.
        """
        try:
            print(f"Creating chart for type: {chart_type}, plot: {plot_name}")
            chart_type = chart_type.lower() if chart_type else 'bar'

            # Initialize chart object
            if chart_type in ['bar', 'column', 'clustered']:
                print(f"Creating BarChart for {plot_name}")
                chart = BarChart()
                chart.type = "col"
                chart.title = plot_name
                chart.x_axis.title = "Categories"
                chart.y_axis.title = "Values"

            elif chart_type == 'line':
                print(f"Creating LineChart for {plot_name}")
                chart = LineChart()
                chart.title = plot_name
                chart.x_axis.title = "Categories"
                chart.y_axis.title = "Values"

            elif chart_type == 'pie':
                print(f"Creating PieChart for {plot_name}")
                chart = PieChart()
                chart.title = plot_name

            elif chart_type == 'scatter':
                print(f"Creating ScatterChart for {plot_name}")
                chart = ScatterChart()
                chart.title = plot_name
                chart.x_axis.title = "X Values"
                chart.y_axis.title = "Y Values"

            elif chart_type == 'stacked':
                print(f"Creating Stacked BarChart for {plot_name}")
                chart = BarChart()
                chart.type = "col"
                chart.grouping = "stacked"
                chart.title = plot_name
                chart.x_axis.title = "Categories"
                chart.y_axis.title = "Values"
                # Enable stacking for proper stacked chart behavior
                chart.overlap = 100  # Ensure bars are stacked on top of each other

            elif chart_type == 'stacked100':
                print(f"Creating Stacked100 BarChart for {plot_name}")
                chart = BarChart()
                chart.type = "col"
                chart.grouping = "percentStacked"
                chart.title = plot_name
                chart.x_axis.title = "Categories"
                chart.y_axis.title = "Percentage (%)"
                # Enable percentage stacking
                chart.overlap = 100  # Ensure bars are stacked on top of each other

            else:
                print(f"Creating default BarChart for {plot_name}")
                chart = BarChart()
                chart.type = "col"
                chart.title = plot_name
                chart.x_axis.title = "Categories"
                chart.y_axis.title = "Values"

            # === Chart appearance enhancements ===
            # Enable legend
            chart.legend = Legend()
            chart.legend.position = "r"  # Right side by default

            # Configure data labels based on chart type
            chart.dLbls = DataLabelList()
            if chart_type == 'stacked100':
                # Show percentage values for stacked 100% charts
                chart.dLbls.showPercent = True
                chart.dLbls.showVal = False
            else:
                # Hide values by default for other chart types
                chart.dLbls.showVal = False

            # === Enable axes by default ===
            # Enable X-axis
            if hasattr(chart, 'x_axis'):
                chart.x_axis.spPr = None  # Ensure X-axis is visible
                chart.x_axis.txPr = None  # Ensure X-axis text is visible
                chart.x_axis.majorGridlines = None  # Disable X-axis gridlines
                chart.x_axis.minorGridlines = None  # Disable X-axis minor gridlines
                # Ensure X-axis is displayed
                chart.x_axis.delete = False
                chart.x_axis.visible = True
            
            # Enable Y-axis
            if hasattr(chart, 'y_axis'):
                chart.y_axis.spPr = None  # Ensure Y-axis is visible
                chart.y_axis.txPr = None  # Ensure Y-axis text is visible
                chart.y_axis.majorGridlines = None  # Disable Y-axis gridlines
                chart.y_axis.minorGridlines = None  # Disable Y-axis minor gridlines
                # Ensure Y-axis is displayed
                chart.y_axis.delete = False
                chart.y_axis.visible = True

            # === Disable gridlines ===
            # Disable major gridlines
            if hasattr(chart, 'majorGridlines'):
                chart.majorGridlines = None
            
            # Disable minor gridlines
            if hasattr(chart, 'minorGridlines'):
                chart.minorGridlines = None

            # Note: Chart size will be set after adding to worksheet
            # The graphicFrame property is only available after the chart is added to a worksheet

            # Disable features that clutter the chart
            if hasattr(chart, 'dispBlanksAs'):
                chart.dispBlanksAs = 'gap'  # Cleaner look
            if hasattr(chart, 'showDLbls'):
                chart.showDLbls = True      # Ensure labels are displayed

            # Disable unnecessary elements (Excel defaults)
            if hasattr(chart, 'displayBlanksAs'):
                chart.displayBlanksAs = 'gap'
            if hasattr(chart, 'layout'):
                chart.layout = None  # Reset layout for simplicity

            print(f"Chart created successfully with enhanced appearance: {type(chart)}")
            return chart

        except Exception as e:
            print(f"Error creating chart for type {chart_type}: {e}")
            print(f"Available chart classes: BarChart={BarChart}, LineChart={LineChart}, PieChart={PieChart}, ScatterChart={ScatterChart}")
            return None


    def _clean_dataframe_for_excel(self, df):
        """
        Clean pandas DataFrame for Excel compatibility by converting Period objects and other non-serializable types.
        """
        try:
            # Create a copy to avoid modifying the original
            df_clean = df.copy()
            
            # First clean column names (remove any problematic characters and Period objects)
            clean_columns = []
            for col in df_clean.columns:
                col_str = str(col)
                # Remove Period_ prefix and clean the column name
                if col_str.startswith('Period_'):
                    col_str = col_str.replace('Period_', '')
                # Clean any other problematic characters
                col_str = col_str.strip().replace('\n', ' ').replace('\r', '').replace('\\', '_').replace('/', '_')
                clean_columns.append(col_str)
            
            df_clean.columns = clean_columns
            
            # Convert each column to handle different data types
            for col in df_clean.columns:
                try:
                    # Handle Period objects first
                    if df_clean[col].dtype == 'object':
                        # Check if this column contains Period objects
                        sample_values = df_clean[col].dropna().head(10)
                        if any(str(val).startswith('Period_') for val in sample_values):
                            # Convert Period objects to strings
                            df_clean[col] = df_clean[col].astype(str).str.replace('Period_', '')
                        else:
                            df_clean[col] = df_clean[col].astype(str)
                    elif 'period' in str(df_clean[col].dtype).lower():
                        # This is a pandas Period dtype
                        df_clean[col] = df_clean[col].astype(str).str.replace('Period_', '')
                    elif 'datetime' in str(df_clean[col].dtype).lower():
                        df_clean[col] = df_clean[col].dt.strftime('%Y-%m-%d %H:%M:%S')
                    elif 'timedelta' in str(df_clean[col].dtype).lower():
                        df_clean[col] = df_clean[col].astype(str)
                    else:
                        # For numeric types, keep as is but ensure no Period objects
                        df_clean[col] = df_clean[col].astype(str)
                except Exception as col_error:
                    print(f"Error processing column {col}: {col_error}")
                    # Fallback: convert to string
                    df_clean[col] = df_clean[col].astype(str)
            
            # Final pass: ensure all values are Excel-compatible
            for col in df_clean.columns:
                try:
                    df_clean[col] = df_clean[col].apply(
                        lambda x: self._safe_convert_value(x)
                    )
                except Exception as val_error:
                    print(f"Error converting values in column {col}: {val_error}")
                    # Fallback: convert everything to string
                    df_clean[col] = df_clean[col].astype(str)
            
            # Additional safety check: ensure no Period objects remain
            for col in df_clean.columns:
                try:
                    # Check if any values still contain Period_ prefix
                    if df_clean[col].astype(str).str.contains('Period_').any():
                        print(f"Warning: Column {col} still contains Period_ prefix, cleaning again")
                        df_clean[col] = df_clean[col].astype(str).str.replace('Period_', '')
                except:
                    pass
            
            return df_clean
            
        except Exception as e:
            print(f"Error cleaning DataFrame for Excel: {e}")
            # Fallback: convert everything to string and clean
            try:
                df_fallback = df.copy()
                # Clean column names
                df_fallback.columns = [str(col).replace('Period_', '').strip() 
                                     for col in df_fallback.columns]
                # Convert all values to strings
                for col in df_fallback.columns:
                    df_fallback[col] = df_fallback[col].astype(str).str.replace('Period_', '')
                return df_fallback
            except:
                # Ultimate fallback
                return pd.DataFrame({'Error': ['Data conversion failed']})

    def _safe_convert_value(self, val):
        """
        Safely convert a value to Excel-compatible format
        """
        try:
            if pd.isna(val) or val is None:
                return ''
            elif hasattr(val, 'to_timestamp'):  # Period objects
                return val.to_timestamp().strftime('%Y-%m-%d')
            elif hasattr(val, 'isoformat'):  # Timestamp objects
                return val.isoformat()
            elif hasattr(val, 'strftime'):  # datetime objects
                return val.strftime('%Y-%m-%d %H:%M:%S')
            elif isinstance(val, (list, dict)):
                return str(val)
            elif isinstance(val, str) and val.startswith('Period_'):
                return val.replace('Period_', '')
            else:
                return str(val)
        except Exception:
            return str(val)

    def _create_clean_test_dataframe(self, plot_name):
        """
        Create a completely clean test DataFrame to isolate Period object issues
        """
        try:
            # Create simple test data with no Period objects
            test_data = [
                {'Category': 'Test 1', 'Value': 100},
                {'Category': 'Test 2', 'Value': 200},
                {'Category': 'Test 3', 'Value': 300},
            ]
            
            df_test = pd.DataFrame(test_data)
            print(f"Created clean test DataFrame for plot {plot_name}")
            return df_test
            
        except Exception as e:
            print(f"Error creating test DataFrame: {e}")
            return pd.DataFrame({'Error': ['Test data creation failed']})

    def _aggressive_clean_dataframe(self, df):
        """
        Aggressively clean DataFrame to remove all Period objects and problematic data types
        """
        try:
            # Create a copy
            df_clean = df.copy()
            
            # Clean column names first
            clean_columns = []
            for col in df_clean.columns:
                col_str = str(col)
                # Remove any Period references
                col_str = col_str.replace('Period_', '').replace('period_', '')
                # Clean other problematic characters
                col_str = col_str.strip().replace('\n', ' ').replace('\r', '').replace('\\', '_').replace('/', '_')
                clean_columns.append(col_str)
            
            df_clean.columns = clean_columns
            
            # Convert all columns to strings to avoid any data type issues
            for col in df_clean.columns:
                try:
                    # Convert to string and clean any Period references
                    df_clean[col] = df_clean[col].astype(str).str.replace('Period_', '').str.replace('period_', '')
                except Exception as e:
                    print(f"Error cleaning column {col}: {e}")
                    # Fallback: try to convert to string anyway
                    df_clean[col] = df_clean[col].astype(str)
            
            return df_clean
            
        except Exception as e:
            print(f"Error in aggressive cleaning: {e}")
            # Ultimate fallback: create a simple DataFrame
            return pd.DataFrame({'Data': ['Cleaning failed']})


            